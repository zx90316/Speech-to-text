"""
SSDLC 檢核表填寫腳本
自動填寫 SSDLC 檢核表 Excel 文件
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import os

# 檢核表路徑
CHECKLIST_FILE = "WI-GA215-附件一、SSDLC檢核表.xlsx"
OUTPUT_FILE = "WI-GA215-附件一、SSDLC檢核表-已填寫.xlsx"

# 檢核結果（是/否/不適用）和說明
CHECKLIST_ANSWERS = {
    # 第一階段：需求分析及規劃階段
    "1.1": ("是", "本系統屬於中低風險應用，已評估數據敏感性並實施對應級別的安全控制"),
    "1.2": ("否", "本系統為輔助性工具，非核心業務系統"),
    "1.3": ("是", "實施郵箱驗證、管理員Token驗證、速率限制及黑名單機制"),
    "1.4": ("是", "驗證碼5分鐘過期，驗證成功後24小時會話過期，連續失敗5次自動封禁30分鐘"),
    "1.5": ("是", "每個用戶使用唯一郵箱地址，所有操作記錄來源IP和用戶ID"),
    "1.6": ("是", "採用郵箱+驗證碼雙因子認證機制"),
    "1.7": ("是", "包含個人資料：郵箱地址、音頻文件，無財務數據"),
    "1.8": ("是", "完整操作日誌系統：auth.log, operation.log, security.log, error.log, audit.log"),
    "1.9": ("否", "當前為單機部署，可擴展至HA架構"),
    "1.10": ("是", "日誌文件自動輪替和備份（保留180-1825天），任務完成後自動清理臨時文件"),

    # 第二階段：架構設計階段
    "2.1": ("是", "資源隔離（每個任務獨立目錄）、異步處理、資源限制（文件500MB，處理4小時）"),
    "2.2": ("是", "每個郵箱地址對應唯一用戶，不允許多人共用同一郵箱"),
    "2.3": ("是", "郵箱+驗證碼雙因子認證，驗證碼6位隨機數字，5分鐘有效期，5次失敗封禁"),
    "2.4": ("是", "驗證碼5分鐘過期，驗證成功24小時會話過期，郵件通知（驗證碼、任務完成），所有登入成功/失敗記錄日誌"),
    "2.5": ("是", "一般日誌保存180天（6個月），涉及個資的audit.log保存1825天（5年）"),
    "2.6": ("是", "所有日誌包含：事件類型、時間戳、IP地址、用戶ID、操作結果、詳細資訊"),
    "2.7": ("是", "日誌輪替機制（單文件10-50MB），超過保存期限自動刪除，異步寫入不阻塞主線程"),
    "2.8": ("是", "日誌寫入失敗會輸出警告，建議生產環境配置監控系統（Prometheus+Grafana）"),
    "2.9": ("部分", "所有日誌使用系統時間，建議容器化部署時配置NTP服務，無UI介面（API服務）"),
    "2.10": ("是", "任務數據：服務器重啟後清除；臨時文件：任務完成/失敗立即刪除；日誌：超過期限自動輪替刪除"),
    "2.11": ("是", "密碼：PBKDF2-SHA256（100,000次迭代）；數據：Fernet加密；郵箱：SHA-256+鹽值；傳輸：HTTPS/TLS 1.2+；日誌遮罩敏感信息"),
    "2.12": ("是", "統一API入口，所有請求經過安全中介軟體，郵箱驗證端點需驗證碼，管理員端點需ADMIN_TOKEN"),
    "2.13": ("是", "HSTS標頭強制HTTPS，支持TLS 1.2+，安全標頭（X-Content-Type-Options, X-Frame-Options, CSP等）"),

    # 第三階段：開發與測試階段
    "3.1": ("是", "已根據SSDLC要求設計和實施安全功能，持續評估和改進"),
    "3.2": ("是", "已文件化：SECURITY.md, SSDLC-COMPLIANCE.md, CLAUDE.md, README.md, .env.example"),
    "3.3": ("是", "開發環境（localhost）、測試環境（Docker容器）、生產環境（獨立部署）通過環境變數區分"),
    "3.4": ("建議", "建議組織部署時分配不同角色：開發人員、安全測試人員、運維人員"),
    "3.5": ("建議", "已實施：輸入驗證、XSS防護、CSRF防護、路徑遍歷防護。建議使用：Bandit, Safety, OWASP ZAP進行測試"),
    "3.6": ("建議", "建議在正式部署前進行功能測試、安全測試、性能測試並製作測試報告"),
    "3.7": ("不適用", "系統測試不需要真實個人資料，可使用模擬音頻和臨時郵箱"),
    "3.8": ("是", "郵箱遮罩（ab***@domain.com）、日誌遮罩、無永久存儲個人資料"),
    "3.9": ("不適用", "使用記憶體存儲，無傳統資料庫，敏感數據使用加密或哈希"),
    "3.10": ("是", "建議使用HTTPS/TLS傳輸，郵件傳輸使用SMTP with STARTTLS，HSTS標頭強制HTTPS"),
    "3.11": ("是", "郵箱驗證（雙因子），管理員Token存放於環境變數，不硬編碼，加密金鑰自動生成"),

    # 第四階段：系統上線階段
    "4.1": ("是", "requirements.txt指定最新穩定版本，使用cryptography >=41.0.0，建議部署前執行pip-audit"),
    "4.2": ("是", "通過環境變數區分環境，不同環境使用不同的.env配置"),
    "4.3": ("建議", "建議組織部署時實施職責分離，不同人員擁有不同的訪問權限"),
    "4.4": ("建議", "建議制定上線計畫：環境準備、依賴安裝、配置設置、安全檢查、服務部署、功能測試、安全測試、監控設置"),
    "4.5": ("建議", "建議使用Git進行版本管理，使用Tags標記版本，維護CHANGELOG.md，當前版本v2.1.0"),
    "4.6": ("建議", "提供完整文檔：README.md, SECURITY.md, CLAUDE.md, API文檔。建議培訓：系統操作、安全配置、日誌監控、應急響應"),

    # 第五階段：維運階段
    "5.1": ("建議", "建議：代碼變更通過Git管理、配置變更記錄在文檔、安全更新優先處理、重大變更需安全審查"),
    "5.2": ("建議", "建議：配置文件(.env)備份、日誌文件定期備份、代碼使用Git版本控制"),
    "5.3": ("建議", "支持Docker容器化部署，支持水平擴展（多實例+負載平衡器）"),
    "5.4": ("是", "任務數據：處理完成後立即刪除；臨時文件：自動清理；日誌：超過期限自動輪替刪除；所有刪除操作記錄在audit.log"),

    # 其他項目
    "6.1": ("是", "日誌：自動輪替備份，保存180-1825天，建議定期備份至外部存儲；數據銷毀：安全刪除機制（覆寫3次），刪除操作記錄在audit.log"),
}

def fill_checklist():
    """填寫檢核表"""
    print("=" * 80)
    print("SSDLC 檢核表自動填寫工具")
    print("=" * 80)

    # 檢查文件是否存在
    if not os.path.exists(CHECKLIST_FILE):
        print(f"[錯誤] 找不到檢核表文件 '{CHECKLIST_FILE}'")
        return

    print(f"\n[讀取] 檢核表：{CHECKLIST_FILE}")

    try:
        # 加載 Excel 文件
        wb = load_workbook(CHECKLIST_FILE)
        ws = wb.active

        print(f"[成功] 加載 Excel 文件，工作表：{ws.title}")
        print(f"[成功] 總行數：{ws.max_row}")

        # 定義顏色
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 淺綠色（是）
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # 淺紅色（否）
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # 淺黃色（不適用/建議）

        # 遍歷所有行，找到檢核項目並填寫
        filled_count = 0

        for row in range(1, ws.max_row + 1):
            # 讀取第一列（編號）
            cell_value = ws.cell(row=row, column=1).value

            if cell_value and str(cell_value).strip() in CHECKLIST_ANSWERS:
                item_id = str(cell_value).strip()
                answer, description = CHECKLIST_ANSWERS[item_id]

                print(f"\n處理項目 {item_id}：{answer}")

                # 根據答案填寫對應的列
                if answer == "是":
                    ws.cell(row=row, column=3).value = "V"  # 第3列：是
                    ws.cell(row=row, column=3).fill = green_fill
                    ws.cell(row=row, column=4).value = ""   # 清空否
                    ws.cell(row=row, column=5).value = ""   # 清空不適用
                elif answer == "否":
                    ws.cell(row=row, column=3).value = ""   # 清空是
                    ws.cell(row=row, column=4).value = "V"  # 第4列：否
                    ws.cell(row=row, column=4).fill = red_fill
                    ws.cell(row=row, column=5).value = ""   # 清空不適用
                elif answer == "不適用":
                    ws.cell(row=row, column=3).value = ""   # 清空是
                    ws.cell(row=row, column=4).value = ""   # 清空否
                    ws.cell(row=row, column=5).value = "V"  # 第5列：不適用
                    ws.cell(row=row, column=5).fill = yellow_fill
                elif answer in ["部分", "建議"]:
                    # 建議項目：在說明中標註
                    ws.cell(row=row, column=3).value = "V"  # 第3列：是（部分）
                    ws.cell(row=row, column=3).fill = yellow_fill
                    ws.cell(row=row, column=4).value = ""   # 清空否
                    ws.cell(row=row, column=5).value = ""   # 清空不適用

                # 填寫說明（第6列）
                ws.cell(row=row, column=6).value = description
                ws.cell(row=row, column=6).font = Font(size=10)

                filled_count += 1
                print(f"  [已填寫] {description[:50]}...")

        # 保存文件
        print(f"\n[保存] 填寫後的檢核表到：{OUTPUT_FILE}")
        wb.save(OUTPUT_FILE)

        print("\n" + "=" * 80)
        print(f"[完成] 共填寫 {filled_count} 個檢核項目")
        print(f"[輸出] 文件：{OUTPUT_FILE}")
        print("=" * 80)

        # 統計結果
        yes_count = sum(1 for answer, _ in CHECKLIST_ANSWERS.values() if answer == "是")
        no_count = sum(1 for answer, _ in CHECKLIST_ANSWERS.values() if answer == "否")
        na_count = sum(1 for answer, _ in CHECKLIST_ANSWERS.values() if answer == "不適用")
        partial_count = sum(1 for answer, _ in CHECKLIST_ANSWERS.values() if answer in ["部分", "建議"])

        print("\n[統計結果]")
        print(f"  [是] {yes_count} 項")
        print(f"  [否] {no_count} 項")
        print(f"  [建議/部分] {partial_count} 項")
        print(f"  [不適用] {na_count} 項")
        print(f"  [總計] {len(CHECKLIST_ANSWERS)} 項")

        compliance_rate = (yes_count + partial_count) / len(CHECKLIST_ANSWERS) * 100
        print(f"\n[合規率] {compliance_rate:.1f}%")

    except Exception as e:
        print(f"\n[錯誤] {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fill_checklist()
